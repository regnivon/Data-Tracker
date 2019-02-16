import tkinter as tk
import openpyxl as xl
from openpyxl.styles import Alignment
from tkinter import messagebox
from collections import OrderedDict
from datetime import datetime

# TODO: Make sure formatting of the excel sheet is maintained through row deletions and additions
# TODO: Implement method to clear empty rows
# TODO: More testing, make look for edge cases in the chronological order sorting
# TODO: Look for better way of handling the day 14 cases besides adding an end to the positions?

# desired fields to build interface from

entry_fields = ["Date", "Paraffin Amount", "Slides Amount", "Day", "Type"]
choices = {"Day": ["1", "4", "8", "14"], "Type": ["Young Saline", "Young Infected", "Age Saline", "Age Infected"]}
rat_positions = OrderedDict({"Young Saline": 2, "Young Infected": 5, "Age Saline": 8, "Age Infected": 11})
day_positions = OrderedDict({"1": 3, "4": 10, "8": 17, "14": 24, "End": 37})

# instantiate GUI object for building
top = tk.Tk()
top.title("Rat Data Tracker")
top.configure(background='pink')

# build the dropbox variables and create a list to store lists of entry fields
entries = []
drop_entries = []
drop1 = tk.StringVar()
drop1.set("Choose Day")
drop2 = tk.StringVar()
drop2.set("Choose Rat Type")
drop_entries.append(drop1)
drop_entries.append(drop2)


# method to write input fields to excel file
def write_fields(entries):
    day = drop_entries[0].get()
    rat = rat_positions[drop_entries[1].get()]
    date = entries[0][1].get()
    if check_date(date, day, rat):
        row = find_date(date, day, rat)
        paraffin = int(sheet.cell(row, rat + 1).value)
        slide = int(sheet.cell(row, rat + 2).value)
        paraffin += int(entries[1][1].get())
        slide += int(entries[2][1].get())
        if paraffin == 0 and slide == 0:
            sheet.cell(row, rat).value = None
            sheet.cell(row, rat+1).value = None
            sheet.cell(row, rat+2).value = None
        else:
            sheet.cell(row, rat + 1).value = paraffin
            sheet.cell(row, rat + 2).value = slide
    else:
        row = find_empty_date(day, rat)
        for entry in entries:
            if entries.index(entry) == 0:
                text = entry[1].get()
                sheet.cell(row, rat, text)
                rat += 1
            else:
                text = entry[1].get()
                sheet.cell(row, rat, text)
                rat += 1
    update_day_dict()
    messagebox.showinfo("Success", "Written to Excel sheet.")


# checks if rat already exists
def check_date(date, day, col):
    days = tuple(day_positions)
    start_row = day_positions[day]
    next_day = days[days.index(day) + 1]
    for row in range(start_row, day_positions[next_day]):
        if sheet.cell(row, col).value == date:
            return True


# finds already existing date
def find_date(date, day, col):
    days = tuple(day_positions)
    start_row = day_positions[day]
    next_day = days[days.index(day) + 1]
    for row in range(start_row, day_positions[next_day]):
        if sheet.cell(row, col).value == date:
            return row


# clears entry fields
def delete_fields():
    for entry in entries:
        entry[1].delete(0, 500)


# finds cell to start putting data, creates a new row if not and updates
# date dictionary
def find_empty_date(day, col):
    days = tuple(day_positions)
    start_row = day_positions[day]
    next_day = days[days.index(day) + 1]
    for cur_row in range(start_row, day_positions[next_day]):
        if sheet.cell(cur_row, col).value is None:
            return cur_row
    sheet.insert_rows(day_positions[next_day])
    update_day_dict()
    return day_positions[next_day]-1


def update_day_dict():
    cur_row = day_positions["1"]
    while sheet.cell(cur_row, 1).value != "Day 14":
        for key in day_positions:
            if key != "End":
                if sheet.cell(cur_row, 1).value == ("Day " + key):
                    day_positions[key] = cur_row
        cur_row += 1
    end = day_positions["14"]
    for rat in rat_positions:
        last_row_in_col = None
        for row in range(day_positions["14"], day_positions["14"] + 15):
            if sheet.cell(row, rat_positions[rat]).value is None:
                last_row_in_col = row
        if last_row_in_col > end:
            end = last_row_in_col
    day_positions["End"] = end


# iterate through days and rat types, checking if the fields are in chronological order.
# Re-write them so they are
def chronological_order():
    date_format = "%m/%d/%Y"
    days = tuple(day_positions)
    for day in choices["Day"]:
        start_row = day_positions[day]
        next_day = days[days.index(day) + 1]
        for rat_type in rat_positions:
            data_list = []
            for row in range(start_row, day_positions[next_day]):
                row_data = []
                if sheet.cell(row, rat_positions[rat_type]).value is not None:
                    row_data.append(str(sheet.cell(row, rat_positions[rat_type]).value))
                    row_data.append(sheet.cell(row, rat_positions[rat_type] + 1).value)
                    row_data.append(sheet.cell(row, rat_positions[rat_type] + 2).value)
                    data_list.append(row_data)
            try:
                data_list = sorted(data_list, key=lambda date: datetime.strptime(date[0][0:10], date_format))
                write_list(data_list, start_row, rat_positions[rat_type])
            except ValueError as er:
                print(er)


def write_list(data_list, start_row, col):
    cur_row = start_row
    for data in data_list:
        sheet.cell(cur_row, col).value = data[0]
        sheet.cell(cur_row, col + 1).value = data[1]
        sheet.cell(cur_row, col + 2).value = data[2]
        cur_row += 1


# builds a sheet if the shared one is not found, otherwise
# just updates the day_positions dictionary
# can make the excel file name read in from a config later if desired
try:
    workbook = xl.load_workbook("Slides.xlsx")
    sheet = workbook.active
    row = 3
    update_day_dict()
except FileNotFoundError:
    row = 3
    col = 2
    workbook = xl.Workbook()
    sheet = workbook.active
    for day in choices["Day"]:
        sheet['A' + str(row)] = "Day " + day
        row += 7
    row = 1
    for rat in choices["Type"]:
        sheet.merge_cells(start_row=row + 1, start_column=col, end_row=row + 1, end_column=col + 2)
        sheet.cell(row + 1, col, rat)
        sheet.cell(row + 1, col).alignment = Alignment(horizontal="center", vertical="center")
        sheet.cell(row, col, "Date")
        col += 1
        sheet.cell(row, col, "# of hippocampus slices")
        col += 1
        sheet.cell(row, col, "# of slides")
        col += 1

# build GUI
for field in entry_fields:
    row = tk.Frame(top)
    label = tk.Label(row, width=20, text=field, anchor='w', fg="gray1", bg="pink")
    row.pack(side="top", fill="x", padx=5, pady=5)
    if entry_fields.index(field) < 3:
        ent = tk.Entry(row)
        ent.pack(side="right", expand="yes", fill="x")
        entries.append((field, ent))
    else:
        menu = tk.OptionMenu(row, drop_entries[entry_fields.index(field)-3], *choices[field])
        menu.pack(side="right", expand="yes", fill="x")
    label.pack(side="left")

entries[0][1].insert(0, "mm/dd/yyyy")
tk.Button(top, text="Quit", command=top.quit, activebackground="red").pack(side="left")
tk.Button(top, text="Submit", command=(lambda e=entries: write_fields(e))).pack(side="left")
tk.Button(top, text="Clear", command=delete_fields).pack(side="left")
top.mainloop()
#delete_empty_rows()
chronological_order()
workbook.save("Slides.xlsx")
